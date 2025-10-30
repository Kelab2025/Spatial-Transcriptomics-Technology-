import sys
import pandas as pd
import re
import xlwt
import xlrd
import openpyxl
from xlutils.copy import copy
from Bio import SeqIO
from Bio.SeqUtils import GC
from Bio.SeqUtils.MeltingTemp import Tm_NN
from Bio.SeqUtils.MeltingTemp import Tm_GC
# from PyQt5 import *
# from PyQt5.QtWidgets import *
from login import *
from PyQt5.QtCore import *
from PyQt5.QtGui import QCursor
# from PyQt5.QtGui import QMouseEvent, QDesktopServices
from search1 import *
from Bio.Blast.Applications import NcbiblastnCommandline


# ###############1.读取maker文件信息到相应的列表
index_start = []  # 探针序列起始位置索引
index_end = []  # 探针序列结束位置索引
Reverse_Complement = []  # 反向互补序列列表
Reverse_Complement_i = []
prober = []
prober_i = []
ATGC = []
numberss = []
lenss = []
es = 0



########## 读取文件
df = pd.read_excel('./2024.xlsx')  # 读取maker文件地址




gene_list = df['Gene'].tolist()  # 获取基因列
# print(gene_list)
# Serial_list = df['Serial_number'].tolist()  # 获取序列号列
# ###############2.读取序列号列表在对应数据库挖掘
#     for gene_tap in gene_list:
#
00#         SEQUENCE_ID = Serial
#     if self.checkBox.isChecked():  # 判断复选框是否被选中
print("HUMAN")
fasta = SeqIO.to_dict(SeqIO.parse("./data/human/rna.fna",'fasta'))
data_excel = xlrd.open_workbook('./data/human/human.xls')
bpcommandline = NcbiblastnCommandline(query="./data/qpcr_primer.fasta",
                                          db="./data/human/data_human/human.fasta", outfmt=6,
                                          out="./output/need_blast.txt")
# 获取book中的sheet工作表的三种方法,返回一个xlrd.sheet.Sheet()对象
table = data_excel.sheet_by_name(sheet_name='Sheet1')  # 通过名称获取
# excel工作表的行列操作
n_rows = table.nrows  # 获取该sheet中的有效行数
n_cols = table.ncols  # 获取该sheet中的有效列数
row_list = table.row(rowx=0)  # 返回某行中所有的单元格对象组成的列表
cols_list = table.col(colx=1)  # 返回某列中所有的单元格对象组成的列表
# 返回某列中所有单元格的数据组成的列表
cols_xulie = table.col_values(0, start_rowx=0, end_rowx=None)
cols_data = table.col_values(1, start_rowx=0, end_rowx=None)
# if self.checkBox_2.isChecked():
#     print("MUS")
#     fasta = SeqIO.to_dict(SeqIO.parse("./data/mouse/GCF_000001635.27_GRCm39_rna.fna",'fasta'))
#     data_excel = xlrd.open_workbook('./data/mouse/mouse.xlsx')
#     bpcommandline = NcbiblastnCommandline(query="./data/qpcr_primer.fasta",
#                                           db="./data/mouse/data_mouse/mouse.fasta", outfmt=6,
#                                           out="./output/need_blast.txt")
#     # 获取book中的sheet工作表的三种方法,返回一个xlrd.sheet.Sheet()对象
#     table = data_excel.sheet_by_name(sheet_name='Sheet1')  # 通过名称获取
#     # excel工作表的行列操作
#     n_rows = table.nrows  # 获取该sheet中的有效行数
#     n_cols = table.ncols  # 获取该sheet中的有效列数
#     row_list = table.row(rowx=0)  # 返回某行中所有的单元格对象组成的列表
#     cols_list = table.col(colx=1)  # 返回某列中所有的单元格对象组成的列表
#     # 返回某列中所有单元格的数据组成的列表
#     cols_xulie = table.col_values(0, start_rowx=0, end_rowx=None)
#     cols_data = table.col_values(1, start_rowx=0, end_rowx=None)
# else:
#     print("Error，未选择或复选基因库")

# ######选择流程顺序
# if self.checkBox_3.isChecked():
#     flow_path = 'Situ Sequencing'
# if self.checkBox_4.isChecked():
#     flow_path = 'Situ Hybridization'
# if self.checkBox_5.isChecked():
#     flow_path = 'PhoL-Fish'


for gene_tap in gene_list:# 基因名列表组
    gene_tapnum = []
    for i in range(len(cols_data)):
        pattern = r'\'(.*?)\''
        matches = re.findall(pattern, cols_data[i])
        for match in matches:
            # print(match)
            if gene_tap == match:
                gene_tapnum.append(i)
            else:
                pass
    gene_tapseq = []
    for i in gene_tapnum:
        col = cols_xulie[i]
        gene_tapseq.append(col)
    gene_tapseqed = []
    print(gene_tapseq)
    for i in gene_tapseq:
        str_delallblank = i.replace(' ', '')
        res = re.findall(r">(.+?)Homo", str_delallblank)
        print(res)
        str1 = ''.join(res)
        gene_tapseqed.append(str1)
    print(gene_tapseqed)
    aaaa = []
    for i in gene_tapseqed:
        if 'NM' in i:
            aaaa.append(i)
        if 'NR' in i:
            aaaa.append(i)
        if 'XM' in i:
            ii = i.rstrip('PREDICTED:')
            aaaa.append(ii)
    print(aaaa)
    gene_lens = []
    for i in aaaa:
        Gene_data = fasta[i]
        Gene_seq = Gene_data.seq
    # with open('output/object.txt', 'w') as f:  # 输出目的基因文件到本地文件夹
    #     f.write(Gene_data)
        gene_len = len(Gene_seq)
        gene_lens.append(gene_len)
    b = gene_lens.index(min(gene_lens))
    len_sequence = min(gene_lens)
    Sequence_id = aaaa[b]
    Gene_data = fasta[Sequence_id]
    Gene_seq = Gene_data.seq
    # print(Sequence_id)

    # ##############3.数据类型转换
    # MIN_SIZE = self.spinBox.text()
    # OPT_SIZE = self.spinBox_2.text()
    # MAX_SIZE = self.spinBox_3.text()
    # MIN_TM = self.doubleSpinBox.text()
    # OPT_TM = self.doubleSpinBox_2.text()
    # MAX_TM = self.doubleSpinBox_3.text()
    # MIN_GC = self.doubleSpinBox_6.text()
    # OPT_GC = self.doubleSpinBox_5.text()
    # MAX_GC = self.doubleSpinBox_4.text()

######## 手动改参数在这里 #############
    ## 探针靶向区域长度 （单边）
    MIN_SIZE = 16
    OPT_SIZE = 16
    MAX_SIZE = 16
    ## 探针TM值范围
    MIN_TM = 40
    OPT_TM = 50
    MAX_TM = 60
    ## 探针GC值范围
    MIN_GC = 40
    OPT_GC = 50
    MAX_GC = 60

    PRIMER_MIN_SIZE = int(MIN_SIZE)
    PRIMER_OPT_SIZE = int(OPT_SIZE)
    PRIMER_MAX_SIZE = int(MAX_SIZE)
    SEQ_length = int(len(Gene_data))
    PRIMER_MIN_TM = float(MIN_TM)
    PRIMER_OPT_TM = float(OPT_TM)
    PRIMER_MAX_TM = float(MAX_TM)
    PRIMER_MIN_GC = float(MIN_GC)
    PRIMER_OPT_GC = float(OPT_GC)
    PRIMER_MAX_GC = float(MAX_GC)

# ##############4.根据ui输入参数值进行探针初步筛选设计
    df_sequences = []  # 储存符合Gc Tm size 条件的序列
    interval = PRIMER_MAX_SIZE - PRIMER_MIN_SIZE
    if interval == 0:  # 当max - min = 0 时 证明探针长度是一定的
        for i in range(SEQ_length - (PRIMER_MIN_SIZE - 1)):
            end = i + PRIMER_OPT_SIZE
            sequence = Gene_seq[i:end]
            GC_sequence = GC(sequence)
            TM_sequence = Tm_GC(sequence)
            if ((PRIMER_MIN_GC <= GC_sequence <= PRIMER_MAX_GC) and (PRIMER_MIN_TM <= TM_sequence <= PRIMER_MAX_TM)) and (len(sequence) == PRIMER_OPT_SIZE):
                df_sequences.append(sequence)
    else:
        for i in range(SEQ_length - (PRIMER_MIN_SIZE - 1)):
            for p in range(interval + 1):
                end = i + PRIMER_MIN_SIZE + p
                sequence = Gene_seq[i:end]
                GC_sequence = GC(sequence)
                TM_sequence = Tm_GC(sequence)
                if ((PRIMER_MIN_GC <= GC_sequence <= PRIMER_MAX_GC) and (PRIMER_MIN_TM <= TM_sequence <= PRIMER_MAX_TM)) and (len(sequence) == 16):
                    df_sequences.append(sequence)
    # pattern = re.compile('([\'])([A-Z])+')

# ##########5.依次输出双链接探针正则表达式匹配的结果
    times = 0  # 循环次数
    total_statified = 0
    inside_xulie = []  # 内部单个基因探针序列存储
    # ############获取第四步筛选出的序列，进行正则匹配获取双链接探针
    for io in df_sequences:
        times = times + 1  # 循环次数
        times = str(times)
        io = str(io)
        pattern_l = re.compile(io + '\w{16}')
        Gene_seq = str(Gene_seq)
        primer_left = pattern_l.search(Gene_seq)
        # ########## 判断探针是否符合要求
        if primer_left == None:
            pass
        else:
            primer_left = primer_left.group()
            # print(primer_left)
            pr1 = primer_left[16:]
            p1_GC = GC(pr1)
            p1_TM = Tm_NN(pr1)
            if ((PRIMER_MIN_GC <= p1_GC <= PRIMER_MAX_GC) and (PRIMER_MIN_TM <= p1_TM <= PRIMER_MAX_TM)):
                total_statified += 1
                inside_xulie.append(primer_left)
                print(inside_xulie)
            else:
                pass
        times = int(times)
# #############6.blast前预处理
    inside_blastxulie = []
    print(len(inside_xulie))
    for j in inside_xulie:
        go = open('data/qpcr_primer.fasta', 'w')
        go.write('>' + gene_list[es] +  '\n'
                 + j + '\n')
        go.close()
# #############7.blast操作
        stdout, stderr = bpcommandline()
# #############8.blast结果处理
        # ##################blast导出的TXT文件转易处理的excel##################### #
        # 读取 txt 文档：防止读取错误，读取时需要指定编码
        l = 40
        fopen = open('./output/need_blast.txt', 'r', encoding='utf-8')
        lines = fopen.readlines()
        # 写入 excel表
        file = openpyxl.Workbook()
        sheet = file.active
        # 新建一个sheet
        sheet.title = "data"
        i = 0
        for line in lines:
            # strip 移出字符串头尾的换行
            line = line.strip('\n')
            # 用','替换掉'\t',很多行都有这个问题，导致不能正确把各个特征值分开
            line = line.replace("\t", ",")
            line = line.split(',')
            # 一共7个字段
            for index in range(len(line)):
                sheet.cell(i + 1, index + 1, line[index])
            # 行数递增
            i = i + 1
        file.save('./output/test_out.xlsx')
        #################################对excel文件的第四行进行百分比计算处理##############################
        worksheet = xlrd.open_workbook('./output/test_out.xlsx')
        wb = copy(worksheet)
        ws = wb.get_sheet(0)
        sheet_names = worksheet.sheet_names()
        for sheet_name in sheet_names:
            sheet = worksheet.sheet_by_name(sheet_name='data')
            rows = sheet.nrows  # 获取行数
            cols = sheet.ncols  # 获取列数，尽管没用到
            cols_per = []
        cols = sheet.col_values(3)
        col_sequence = sheet.col_values(1)

        def format_percentage(a, b):
            p = 100 * a / b
            return p

        for w in cols:
            w = int(w)
            q = format_percentage(w, l)
            cols_per.append(q)
        for i in range(len(cols_per)):
            ws.write(i, 3, cols_per[i])
        wb.save('./output/test_out.xlsx')
        level = 0
        for s in col_sequence:
            if s in aaaa:
                pass
            else:
                level = 1
        if level == 0:
            inside_blastxulie.append(j)
            prober.append(j)


    # #############9. 双链接序列反向互补
    print(len(inside_blastxulie))
    for seq in inside_blastxulie:
        seqreverse = seq[::-1]
        transtable = str.maketrans('ATGC', 'TACG')
        finalseq = seqreverse.translate(transtable)
        Reverse_Complement.append(finalseq)
        numberss.append(gene_list[es])
        lenss.append(len_sequence)
# #############10. 获取序列的位置信息
    for a in inside_blastxulie:
        Gene_seq = str(Gene_seq)
        i = Gene_seq.index(str(a)) + 1
        j = i + PRIMER_OPT_SIZE * 2 - 1
        index_start.append(i)
        index_end.append(j)
    es = es + 1
# #############11.最终结果保存输出
for i in prober:
    tr_i = list(i)
    tr_i.insert(16," ")
    ourstr = "".join(tr_i)
    basic = i[15:17]
    ATGC.append(basic)
    prober_i.append(ourstr)

for i in Reverse_Complement:
    tr_i = list(i)
    tr_i.insert(16, " ")
    ourstr = "".join(tr_i)
    Reverse_Complement_i.append(ourstr)
gene_na = 'output'
l = xlwt.Workbook('encoding = utf-8')  # 设置工作簿编码
sheet1 = l.add_sheet('sheet1', cell_overwrite_ok=True)  # 创建sheet工作表
sheet1.write(0, 0, 'Probe')
sheet1.write(0, 1, 'Sequence')
sheet1.write(0, 2, 'Reverse_Complementary')
sheet1.write(0, 3, 'index_start')
sheet1.write(0, 4, 'index_end')
sheet1.write(0, 5, 'basic_group')
sheet1.write(0, 6, 'len')
for i in range(len(numberss)):
    sheet1.write(i + 1, 0, numberss[i])  # 写入数据参数对应 行, 列, 值
for i in range(len(prober_i)):
    sheet1.write(i + 1, 1, prober_i[i])  # 写入数据参数对应 行, 列, 值
for i in range(len(Reverse_Complement_i)):
    sheet1.write(i + 1, 2, Reverse_Complement_i[i])  # 写入数据参数对应 行, 列, 值
for i in range(len(index_start)):
    sheet1.write(i + 1, 3, index_start[i])  # 写入数据参数对应 行, 列, 值
for i in range(len(index_end)):
    sheet1.write(i + 1, 4, index_end[i])  # 写入数据参数对应 行, 列, 值
for i in range(len(ATGC)):
    sheet1.write(i + 1, 5, ATGC[i])  # 写入数据参数对应 行, 列, 值
for i in range(len(lenss)):
    sheet1.write(i + 1, 6, lenss[i])  # 写入数据参数对应 行, 列, 值

file_path = './output/' + gene_na + '.xlsx'  ##输出文件路径
l.save(file_path)  # 保存.xls到当前工作目录
print('探针序列设计完成')